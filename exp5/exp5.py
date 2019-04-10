from openpyxl import load_workbook, Workbook


if __name__ == '__main__':
    wb = load_workbook('exp5_data.xlsx')
    ws = wb.active

    wb2 = Workbook()
    ws2 = wb2.active

    emission = list()
    absorbance = list()

    for c in range(2, 1202):
        pair = [
            round(ws['E' + str(c)].value, 1),  # E for HCl, G for NaOH
            ws['F' + str(c)].value  # F for HCl, H for NaOH
        ]
        emission.append(pair)

    e_max = max(emission, key=lambda l: l[1])[1]
    print(max(emission, key=lambda l: l[1]))
    for e in emission:
        e[1] /= e_max

    for c in range(2, 103):
        pair = [
            round(ws['R' + str(c)].value, 1),
            ws['S' + str(c)].value,  # S for HCl, W for NaOH
            None
        ]
        absorbance.append(pair)

    a_max = max(absorbance, key=lambda l: l[1])[1]
    print(max(absorbance, key=lambda l: l[1]))
    for a in absorbance:
        a[1] /= a_max

    while len(emission):
        inserted = False
        a = 0
        while a < len(absorbance):
            if emission[0][0] == absorbance[a][0]:
                absorbance[a][2] = emission[0][1]
                emission.pop(0)
                inserted = True
                break
            elif absorbance[a][0] > emission[0][0]:
                absorbance.insert(a, [emission[0][0], None, emission[0][1]])
                emission.pop(0)
                inserted = True
                break
            a += 1
        if not inserted:
            absorbance.append([emission[0][0], None, emission[0][1]])
            emission.pop(0)

    for line in absorbance:
        ws2.append(line)
    wb2.save('NaOH.xlsx')
